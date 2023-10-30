import openpyxl


def getRowCount(path, sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColCount(path, sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_column


def getCellData(path, sheetName, rowNumber, colNumber):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.cell(row=rowNumber, column=colNumber).value


def setCellData(path, sheetName, rowNumber, colNumber, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    sheet.cell(row=rowNumber, column=colNumber).value = data
    workbook.save('./new.xlsx')


path = "sample.xlsx"
sheetName = "Quiz"
total_rows = getRowCount(path, sheetName)
total_cols = getColCount(path, sheetName)
print(f'total rows are: {total_rows}, and total cols are: {total_cols}')

print(getCellData(path, sheetName, 2, 2))

# Read
for rows in range(1, total_rows + 1):
    for cols in range(1, total_cols + 1):
        print(getCellData(path, sheetName, rows, cols))

# Write
for rows in range(14, 16):
    for cols in range(1, 5):
        setCellData(path, sheetName, rows, cols, 'Hello')
